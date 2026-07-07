import json
from pathlib import Path
from datetime import datetime, date, time

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]

INPUT_EXCEL = BASE_DIR / "data" / "question_master.xlsx"
OUTPUT_JSON = BASE_DIR / "public" / "data" / "questions.json"
SHEET_NAME = "問題マスター"


REQUIRED_COLUMNS = [
    "id",
    "source",
    "sourceQuestionNo",
    "type",
    "category",
    "subCategory",
    "tags",
    "question",
    "imageFile",
    "choiceA",
    "choiceB",
    "choiceC",
    "choiceD",
    "choiceAExplanation",
    "choiceBExplanation",
    "choiceCExplanation",
    "choiceDExplanation",
    "answer",
    "explanation",
    "explanationImageFile",
    "difficulty",
    "isCalculation",
    "shuffleChoices",
    "isActive",
    "version",
    "lastUpdated",
]


def normalize_value(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")

    if isinstance(value, time):
        return value.strftime("%H:%M:%S")

    return str(value)


def to_bool(value, default=False):
    value = normalize_value(value)

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return value != 0

    if isinstance(value, str):
        text = value.strip().lower()

        if text in ["true", "t", "yes", "y", "1", "○", "〇"]:
            return True

        if text in ["false", "f", "no", "n", "0", "×", "x"]:
            return False

    return default


def split_tags(value):
    value = normalize_value(value)

    if not value:
        return []

    text = str(value).replace("、", ",").replace("，", ",")
    return [tag.strip() for tag in text.split(",") if tag.strip()]


def build_image_path(value):
    value = normalize_value(value)

    if not value:
        return None

    file_name = str(value).strip()

    # Excel側では i-ks-001 / e-ks-001 のように拡張子なしで入力する想定。
    # 拡張子がない場合は .png を補完する。
    if "." not in file_name:
        file_name = f"{file_name}.png"

    return f"/images/{file_name}"


def get_cell(row_dict, column_name):
    return normalize_value(row_dict.get(column_name, ""))


def validate_headers(headers):
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]

    if missing:
        raise ValueError("Excelに必要な列がありません: " + ", ".join(missing))


def build_true_false_question(row):
    raw_answer = get_cell(row, "answer")
    answer_text = str(raw_answer).strip().lower()

    if answer_text not in ["true", "false"]:
        raise ValueError(
            f"{get_cell(row, 'id')}: true_false問題のanswerは true または false にしてください。"
        )

    answer_bool = answer_text == "true"

    return {
        "id": get_cell(row, "id"),
        "source": get_cell(row, "source"),
        "sourceQuestionNo": str(get_cell(row, "sourceQuestionNo")),
        "type": "true_false",
        "category": get_cell(row, "category") or None,
        "subCategory": get_cell(row, "subCategory") or None,
        "tags": split_tags(get_cell(row, "tags")),
        "question": get_cell(row, "question"),
        "image": build_image_path(get_cell(row, "imageFile")),
        "choices": None,
        "answer": answer_bool,
        "explanation": get_cell(row, "explanation"),
        "explanationImage": build_image_path(get_cell(row, "explanationImageFile")),
        "difficulty": get_cell(row, "difficulty") or "normal",
        "isCalculation": to_bool(get_cell(row, "isCalculation")),
        "shuffleChoices": False,
        "isActive": True,
        "version": int(get_cell(row, "version") or 1),
        "lastUpdated": get_cell(row, "lastUpdated"),
    }


def build_multiple_choice_question(row):
    answer = str(get_cell(row, "answer")).strip().upper()

    if answer not in ["A", "B", "C", "D"]:
        raise ValueError(
            f"{get_cell(row, 'id')}: multiple_choice問題のanswerは A/B/C/D のいずれかにしてください。"
        )

    choices = []

    for key in ["A", "B", "C", "D"]:
        choice_text = get_cell(row, f"choice{key}")
        choice_explanation = get_cell(row, f"choice{key}Explanation")

        if not choice_text:
            raise ValueError(
                f"{get_cell(row, 'id')}: choice{key} が空欄です。択一問題ではA～Dをすべて入力してください。"
            )

        choices.append(
            {
                "id": key,
                "text": choice_text,
                "isCorrect": key == answer,
                "explanation": choice_explanation,
            }
        )

    return {
        "id": get_cell(row, "id"),
        "source": get_cell(row, "source"),
        "sourceQuestionNo": str(get_cell(row, "sourceQuestionNo")),
        "type": "multiple_choice",
        "category": get_cell(row, "category") or None,
        "subCategory": get_cell(row, "subCategory") or None,
        "tags": split_tags(get_cell(row, "tags")),
        "question": get_cell(row, "question"),
        "image": build_image_path(get_cell(row, "imageFile")),
        "choices": choices,
        "answer": answer,
        "explanation": get_cell(row, "explanation"),
        "explanationImage": build_image_path(get_cell(row, "explanationImageFile")),
        "difficulty": get_cell(row, "difficulty") or "normal",
        "isCalculation": to_bool(get_cell(row, "isCalculation")),
        "shuffleChoices": to_bool(get_cell(row, "shuffleChoices"), default=True),
        "isActive": True,
        "version": int(get_cell(row, "version") or 1),
        "lastUpdated": get_cell(row, "lastUpdated"),
    }


def validate_question_common(row):
    question_id = get_cell(row, "id")
    question_type = get_cell(row, "type")
    question_text = get_cell(row, "question")

    if not question_id:
        raise ValueError("id が空欄の行があります。")

    if question_type not in ["true_false", "multiple_choice"]:
        raise ValueError(
            f"{question_id}: type は true_false または multiple_choice にしてください。"
        )

    if not question_text:
        raise ValueError(f"{question_id}: question が空欄です。")


def main():
    if not INPUT_EXCEL.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {INPUT_EXCEL}")

    wb = load_workbook(INPUT_EXCEL, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SHEET_NAME}")

    ws = wb[SHEET_NAME]

    headers = []
    for cell in ws[1]:
        headers.append(normalize_value(cell.value))

    validate_headers(headers)

    questions = []
    seen_ids = set()

    for row_number, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        row = {
            headers[index]: normalize_value(cell.value)
            for index, cell in enumerate(row_cells)
            if index < len(headers)
        }

        # 完全空行はスキップ
        if not any(row.values()):
            continue

        # isActive が FALSE の行は出力しない
        if not to_bool(get_cell(row, "isActive"), default=False):
            continue

        try:
            validate_question_common(row)

            question_id = get_cell(row, "id")

            if question_id in seen_ids:
                raise ValueError(f"id が重複しています: {question_id}")

            seen_ids.add(question_id)

            question_type = get_cell(row, "type")

            if question_type == "true_false":
                question = build_true_false_question(row)
            elif question_type == "multiple_choice":
                question = build_multiple_choice_question(row)
            else:
                raise ValueError(f"{question_id}: 未対応のtypeです: {question_type}")

            questions.append(question)

        except Exception as exc:
            raise ValueError(f"Excel {row_number} 行目でエラー: {exc}") from exc

    output = {
        "schemaVersion": 1,
        "generatedAt": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
        "questionCount": len(questions),
        "questions": questions,
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"questions.json を生成しました: {OUTPUT_JSON}")
    print(f"出力問題数: {len(questions)}")


if __name__ == "__main__":
    main()